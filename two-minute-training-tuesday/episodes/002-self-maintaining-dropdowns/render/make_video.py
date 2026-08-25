"""Render Two-Minute Training Tuesday episode 002 to an MP4.

    pip install openpyxl playwright
    python make_video.py                 # full render
    python make_video.py --probe 0,20,30 # just those timestamps, as PNGs, to eyeball

Reads the real rows out of demo-workbook.xlsx, injects them into episode.html,
walks the timeline frame by frame pulling each frame straight off the canvas,
then hands the sequence to ffmpeg. The renderer is deterministic, so the same
inputs always produce the same video.

Needs an ffmpeg with libx264 (the one bundled with Playwright is VP8-only).
"""

import argparse
import base64
import json
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

HERE = Path(__file__).resolve().parent
EPISODE_HTML = HERE / "episode.html"
WORKBOOK = HERE.parent.parent / "001-excel-dynamic-arrays" / "demo-workbook-solved.xlsx"
BRAND_JSON = HERE.parent.parent.parent / "brand.json"
OUT_MP4 = HERE.parent / "TMTT-002-Data-Validation.mp4"
OUT_MP4_CLEAN = HERE.parent / "TMTT-002-Data-Validation-no-captions.mp4"

FPS = 20
WIDTH, HEIGHT = 1920, 1080
# Playwright ships a stripped ffmpeg (VP8 only), so prefer a real one on PATH.
FFMPEG = shutil.which("ffmpeg") or "/opt/pw-browsers/ffmpeg-1011/ffmpeg-linux"
CHROME = "/opt/pw-browsers/chromium-1194/chrome-linux/chrome"


def read_rows():
    ws = load_workbook(WORKBOOK)["ERP Export"]
    return [
        [r[0].strftime("%m/%d/%Y"), r[1], r[2], r[3], r[4], round(r[5], 2)]
        for r in ws.iter_rows(min_row=2, values_only=True)
    ]


def read_brand():
    """Brand colors from brand.json. Missing file just means the built-in defaults."""
    if not BRAND_JSON.exists():
        return {}
    raw = json.loads(BRAND_JSON.read_text())
    return {k: v for k, v in raw.items() if not k.startswith("_")}


def open_page(pw, rows, brand, captions=True):
    browser = pw.chromium.launch(
        executable_path=CHROME if Path(CHROME).exists() else None,
        args=["--force-device-scale-factor=1", "--hide-scrollbars"],
    )
    page = browser.new_page(viewport={"width": WIDTH, "height": HEIGHT})
    page.add_init_script(
        f"window.__ROWS__ = {json.dumps(rows)};"
        f"window.__BRAND__ = {json.dumps(brand)};"
        f"window.__NO_CAPTIONS__ = {json.dumps(not captions)};"
    )
    page.goto(EPISODE_HTML.as_uri())
    page.wait_for_function("window.drawFrame !== undefined")
    return browser, page


def grab(page, t, fmt="jpeg", quality=0.94):
    """Draw frame t and return its bytes, straight off the canvas."""
    page.evaluate("t => window.drawFrame(t)", t)
    data = page.evaluate(
        "([fmt, q]) => document.getElementById('c').toDataURL('image/' + fmt, q).split(',')[1]",
        [fmt, quality],
    )
    return base64.b64decode(data)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--probe", help="comma-separated timestamps to render as PNGs instead")
    ap.add_argument("--outdir", default=None, help="where probe PNGs land")
    ap.add_argument("--no-captions", action="store_true",
                    help="render a clean picture, for adding your own captions in Clipchamp")
    args = ap.parse_args()

    rows = read_rows()
    brand = read_brand()
    print(f"{len(rows)} rows from {WORKBOOK.name}")
    print(f"brand: {brand or 'built-in defaults'}")

    with sync_playwright() as pw:
        browser, page = open_page(pw, rows, brand, captions=not args.no_captions)
        duration = page.evaluate("window.EPISODE_END")

        if args.probe:
            outdir = Path(args.outdir or HERE)
            outdir.mkdir(parents=True, exist_ok=True)
            for spec in args.probe.split(","):
                t = float(spec)
                path = outdir / f"probe-{t:06.2f}.png"
                path.write_bytes(grab(page, t, fmt="png"))
                print("wrote", path)
            browser.close()
            return

        frames_dir = Path(args.outdir or (HERE / "_frames"))
        if frames_dir.exists():
            shutil.rmtree(frames_dir)
        frames_dir.mkdir(parents=True)

        total = int(round(duration * FPS))
        print(f"rendering {total} frames ({duration:.1f}s at {FPS}fps)")
        for i in range(total):
            (frames_dir / f"f{i:06d}.jpg").write_bytes(grab(page, i / FPS))
            if i % 200 == 0:
                print(f"  {i}/{total}", flush=True)
        browser.close()

    out_path = OUT_MP4_CLEAN if args.no_captions else OUT_MP4
    print("encoding with", FFMPEG)
    subprocess.run(
        [
            FFMPEG, "-y", "-hide_banner", "-loglevel", "error",
            "-framerate", str(FPS),
            "-i", str(frames_dir / "f%06d.jpg"),
            "-c:v", "libx264",
            "-preset", "slow",
            "-crf", "20",
            "-pix_fmt", "yuv420p",
            # Even dimensions and faststart so it streams in Teams and SharePoint.
            "-movflags", "+faststart",
            str(out_path),
        ],
        check=True,
    )
    shutil.rmtree(frames_dir)
    size_mb = out_path.stat().st_size / 1e6
    print(f"wrote {out_path} ({size_mb:.1f} MB, {duration:.1f}s)")


if __name__ == "__main__":
    sys.exit(main())
