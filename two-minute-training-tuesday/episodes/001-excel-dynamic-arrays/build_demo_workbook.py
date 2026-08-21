"""Build the on-screen demo workbook for Two-Minute Training Tuesday episode 001.

The workbook stands in for a job-cost export out of the ERP so the episode can
be recorded without a single row of real customer or vendor data on screen.

    pip install openpyxl
    python build_demo_workbook.py

Output: demo-workbook.xlsx, next to this script.

Sheets
------
ERP Export   ~380 rows in a table named ERP_Export. This is the "raw export".
Report       The blank staging area the presenter types the three formulas into.
Answer Key   The finished formulas, for rehearsal and for anyone who wants to
             reuse the file. Keep this sheet off screen while recording.

Data is generated from a fixed seed, so re-running produces a byte-identical
workbook and the script's timecodes stay accurate.
"""

import json
import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

SEED = 20260825
ROW_COUNT = 380
START_DATE = date(2026, 6, 1)
DAY_SPAN = 74

COST_CENTERS = [
    "Powder Line 1",
    "Liquid Line 2",
    "Prep & Blast",
    "Maintenance",
    "QC Lab",
    "Shipping",
]

# Vendor -> the categories that vendor plausibly bills for, so the demo data
# survives someone actually reading it on screen.
VENDORS = {
    "Sherwin-Williams": ["Liquid Coating", "Consumables"],
    "PPG Industries": ["Liquid Coating", "Powder Coating"],
    "Axalta Coating Systems": ["Powder Coating", "Liquid Coating"],
    "Tiger Drylac": ["Powder Coating"],
    "IFS Coatings": ["Powder Coating"],
    "Cardinal Paint": ["Powder Coating"],
    "Graco": ["Equipment Parts", "Tooling"],
    "Nordson": ["Equipment Parts", "Tooling"],
    "Motion Industries": ["Equipment Parts"],
    "Grainger": ["Consumables", "PPE", "Tooling"],
    "Fastenal": ["Consumables", "PPE"],
    "Uline": ["Consumables", "PPE"],
    "Clean Harbors": ["Chemicals"],
    "Old Dominion Freight": ["Freight"],
    "R+L Carriers": ["Freight"],
}

# Category -> the cost centers that plausibly get charged for it. Without this
# the export puts powder purchases on the Shipping cost center, and anyone who
# actually reads the screen stops trusting the demo.
CATEGORY_COST_CENTERS = {
    "Powder Coating": ["Powder Line 1"],
    "Liquid Coating": ["Liquid Line 2"],
    "Consumables": ["Powder Line 1", "Liquid Line 2", "Prep & Blast", "Shipping"],
    "Equipment Parts": ["Powder Line 1", "Liquid Line 2", "Maintenance"],
    "Freight": ["Shipping"],
    "PPE": ["Powder Line 1", "Liquid Line 2", "Prep & Blast", "Maintenance", "QC Lab"],
    "Chemicals": ["Prep & Blast", "QC Lab", "Maintenance"],
    "Tooling": ["Prep & Blast", "Maintenance", "Powder Line 1"],
}

# Category -> (low, high) dollar range.
AMOUNT_RANGE = {
    "Powder Coating": (420, 8600),
    "Liquid Coating": (310, 7400),
    "Consumables": (35, 940),
    "Equipment Parts": (120, 5200),
    "Freight": (85, 1650),
    "PPE": (40, 620),
    "Chemicals": (260, 3100),
    "Tooling": (75, 2400),
}

HEADERS = ["Date", "Job Number", "Cost Center", "Vendor", "Category", "Amount"]

def _brand():
    """Brand colors from ../../brand.json, so the workbook matches the video."""
    path = Path(__file__).resolve().parents[2] / "brand.json"
    defaults = {"header": "#a8121f", "tint": "#fdecee", "input": "#fff0dc"}
    if path.exists():
        raw = json.loads(path.read_text())
        defaults.update({k: v for k, v in raw.items() if k in defaults})
    return {k: v.lstrip("#").upper() for k, v in defaults.items()}


BRAND = _brand()

HEADER_FILL = PatternFill("solid", fgColor=BRAND["header"])
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LABEL_FONT = Font(bold=True, size=11, color=BRAND["header"])
TITLE_FONT = Font(bold=True, size=16, color=BRAND["header"])
HINT_FONT = Font(italic=True, size=10, color="666666")
INPUT_FILL = PatternFill("solid", fgColor=BRAND["input"])
THIN = Side(style="thin", color="BFBFBF")


def build_rows(rng):
    """Generate the fake ERP export rows."""
    vendor_names = list(VENDORS)
    # A handful of jobs, reused across many lines, the way a real export looks.
    jobs = [f"J-26{rng.randint(100, 999)}" for _ in range(28)]
    rows = []
    for _ in range(ROW_COUNT):
        vendor = rng.choice(vendor_names)
        category = rng.choice(VENDORS[vendor])
        low, high = AMOUNT_RANGE[category]
        rows.append(
            [
                START_DATE + timedelta(days=rng.randint(0, DAY_SPAN)),
                rng.choice(jobs),
                rng.choice(CATEGORY_COST_CENTERS[category]),
                vendor,
                category,
                round(rng.uniform(low, high), 2),
            ]
        )
    rows.sort(key=lambda r: r[0])
    return rows


def write_export_sheet(ws, rows):
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    last_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=last_row):
        row[0].number_format = "mm/dd/yyyy"
        row[5].number_format = '$#,##0.00'
        for cell in row:
            cell.border = Border(bottom=THIN)

    # Wide enough to read at 130-150% zoom on camera.
    for col, width in zip("ABCDEF", (13, 14, 17, 26, 19, 14)):
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 22

    table = Table(displayName="ERP_Export", ref=f"A1:F{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"
    return last_row


def write_report_sheet(ws):
    ws["B1"] = "Vendor spend report"
    ws["B1"].font = TITLE_FONT
    ws["B2"] = "Type the three formulas below. Nothing here is hard-coded."
    ws["B2"].font = HINT_FONT

    ws["B4"] = "1. Every vendor we used"
    ws["D4"] = "2. Same list, alphabetical"
    ws["F4"] = "Cost center"
    for ref in ("B4", "D4", "F4"):
        ws[ref].font = LABEL_FONT

    ws["G4"] = COST_CENTERS[0]
    ws["G4"].fill = INPUT_FILL
    ws["G4"].font = Font(bold=True)
    ws["G4"].border = Border(bottom=THIN, top=THIN, left=THIN, right=THIN)

    ws["F6"] = "3. Every line for that cost center"
    ws["F6"].font = LABEL_FONT

    picker = DataValidation(
        type="list",
        formula1='"{}"'.format(",".join(COST_CENTERS)),
        allow_blank=False,
        showDropDown=False,  # False here means "show the in-cell dropdown arrow"
    )
    picker.prompt = "Pick a cost center"
    picker.promptTitle = "Cost center"
    ws.add_data_validation(picker)
    picker.add(ws["G4"])

    for col, width in zip("ABCDEFGHIJKL", (3, 26, 3, 26, 3, 15, 17, 19, 26, 20, 15, 15)):
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


def write_answer_key(ws, last_row):
    ws["B1"] = "Answer key — keep this sheet off screen while recording"
    ws["B1"].font = TITLE_FONT
    ws["B2"] = (
        f"Source: 'ERP Export'!A1:F{last_row}, a table named ERP_Export. "
        "The reference list below is deliberately plain text so it cannot spill."
    )
    ws["B2"].font = HINT_FONT

    ws["B4"] = "Step"
    ws["C4"] = "Formula  (type the = yourself)"
    for ref in ("B4", "C4"):
        ws[ref].font = HEADER_FONT
        ws[ref].fill = HEADER_FILL

    steps = [
        ("1. Unique vendors", "UNIQUE(ERP_Export[Vendor])"),
        ("2. Sorted", "SORT(UNIQUE(ERP_Export[Vendor]))"),
        (
            "3. Filtered by cost center",
            'FILTER(ERP_Export,ERP_Export[Cost Center]=Report!G4,"No matches")',
        ),
        (
            "Bonus: biggest amount first",
            'SORT(FILTER(ERP_Export,ERP_Export[Cost Center]=Report!G4,"No matches"),6,-1)',
        ),
    ]
    for offset, (label, formula) in enumerate(steps):
        row = 5 + offset
        ws.cell(row=row, column=2, value=label).font = LABEL_FONT
        cell = ws.cell(row=row, column=3, value=formula)
        cell.font = Font(name="Consolas", size=10)
        cell.alignment = Alignment(vertical="center")

    # Live versions, placed where their spill ranges cannot collide: the vendor
    # list spills down column B only, the filter spills across F:K.
    ws["B11"] = "Live check — sorted vendors"
    ws["B11"].font = LABEL_FONT
    ws["B12"] = "=SORT(UNIQUE(ERP_Export[Vendor]))"
    ws["D11"] = "distinct vendors"
    ws["D11"].font = HINT_FONT
    ws["D12"] = "=ROWS(UNIQUE(ERP_Export[Vendor]))"
    ws["D12"].font = Font(bold=True, size=14)

    ws["F11"] = "Live check — Report!G4 cost center, biggest amount first"
    ws["F11"].font = LABEL_FONT
    ws["F12"] = (
        '=SORT(FILTER(ERP_Export,ERP_Export[Cost Center]=Report!G4,"No matches"),6,-1)'
    )

    for col, width in zip("ABCDEFGHIJK", (3, 30, 74, 18, 3, 13, 14, 17, 26, 19, 14)):
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


def main():
    rng = random.Random(SEED)
    rows = build_rows(rng)

    wb = Workbook()
    export = wb.active
    export.title = "ERP Export"
    last_row = write_export_sheet(export, rows)

    write_report_sheet(wb.create_sheet("Report"))
    write_answer_key(wb.create_sheet("Answer Key"), last_row)

    wb.active = 0
    out = Path(__file__).with_name("demo-workbook.xlsx")
    wb.save(out)

    vendors = sorted({r[3] for r in rows})
    print(f"Wrote {out}")
    print(f"  {len(rows)} rows, {len(vendors)} distinct vendors, "
          f"{len(COST_CENTERS)} cost centers")
    print(f"  table ERP_Export = 'ERP Export'!A1:F{last_row}")


if __name__ == "__main__":
    main()
